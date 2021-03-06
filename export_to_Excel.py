#Export the Random Walk data to an Excel

#Important Modules
from openpyxl import Workbook

class ExpExcel:
	
	def __init__(self, coord_x, coord_y, iterator):
		
		self.coord_x = coord_x
		self.coord_y = coord_y
		self.iterator = iterator

	def create_excel_files(self): 

		wb = Workbook()

		newexcel_name = f"File number {self.iterator}.xlsx"

		f1 = wb.active

		f1.cell(1,1).value = "x coordinates"
		f1.cell(1,2).value = "y coordinates"

		j = 2

		for x_values in self.coord_x:
			for y_values in self.coord_y:
				f1.cell(j,1).value = x_values
				f1.cell(j,2).value = y_values
				j += 1

		wb.save(newexcel_name)
		wb.close()









